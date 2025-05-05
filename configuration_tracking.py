import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

# ===== TITLE =====
st.markdown(
    "<h1 style='text-align: center; color: #4B0082;'>Configuration Tracking File Generator!!</h1>",
    unsafe_allow_html=True
)
st.markdown(
    "<p style='text-align: center; color: #4682B4;'>Upload the required files, click 'Run Processing', and download the updated file.</p>",
    unsafe_allow_html=True
)

# ===== FILE UPLOADS =====
prev_file = st.file_uploader("Upload previous configuration file", type=["xlsx"])
argo_file = st.file_uploader("Upload Argo file", type=["xlsx"])
vbac_file = st.file_uploader("Upload VBAC/VBAP file", type=["xlsx"])
sales_file = st.file_uploader("Upload Sales file", type=["xlsx"])

if prev_file and argo_file and vbac_file and sales_file:
    st.success("All files uploaded successfully!")

    if st.button("Run Processing"):
        # Load files
        prev_excel = pd.ExcelFile(prev_file)
        prev = pd.read_excel(prev_excel, sheet_name='Configuration tracking')
        ct = pd.read_excel(prev_excel, sheet_name='CT')
        argo = pd.read_excel(argo_file)
        vbac = pd.read_excel(vbac_file)
        sales = pd.read_excel(sales_file)

        # Preprocessing
        prev['CT'] = ""
        ct['Build Product'] = ct['Build Product'].str.upper()
        argo = argo[argo['Division'] == 'PCB']
        argo = argo[argo['Build Product'].isin(ct['Build Product'])]
        argo = argo[~argo['Ship Revenue Type'].isin(['Buyout', 'Qtrly Financial'])]

        vbac = vbac.rename(columns={
            'Sales Doc.': 'Sales Order',
            'Item Forecast ID': 'Forecast ID',
            'Created on': 'SO date'
        })
        sales = sales.rename(columns={'PO Date': 'PO expected date'})

        vbac_unique_so = vbac.drop_duplicates(subset=['Sales Order'])
        vbac_unique_id = vbac.drop_duplicates(subset=['Forecast ID'])

        main_df = pd.merge(argo, vbac_unique_so[['Sales Order', 'SO date']], on='Sales Order', how='left')
        main_df = pd.merge(main_df, vbac_unique_id[['Forecast ID', 'PO date']], on='Forecast ID', how='left')
        main_df = pd.merge(main_df, sales[['Argo ID', 'PO expected date']], on='Argo ID', how='left')

        main_df = pd.merge(main_df, prev[['Argo ID', 'Gate 2.7 actual', 'Gate 3.5 actual', 'Gate 6.5 actual', 'Gate 5.5 actual']], on='Argo ID', how='left')

        for col in ['Gate 2.7 plan', 'Gate 3.5 plan', 'target date Gate 5.5 plan', 'Gate 5.5 plan', 'Gate 6.5 plan']:
            main_df[col] = ""

        # Merge & update CT
        filtered_prev_pp = prev[~prev['Argo ID'].isin(main_df['Argo ID'])]
        main_df['CT'] = ""
        matching_rows_prev_pp = prev[prev['Argo ID'].isin(main_df['Argo ID'])]
        main_df.loc[main_df['Argo ID'].isin(prev['Argo ID']), 'CT'] = main_df['Argo ID'].map(
            matching_rows_prev_pp.set_index('Argo ID')['CT']
        )

        main_df = pd.concat([main_df, filtered_prev_pp], ignore_index=True)

        main_df['Build Product'] = main_df['Build Product'].astype(str)
        ct = ct.rename(columns={'CT': 'CT_temp'})
        main_df = pd.merge(main_df, ct, on='Build Product', how='left')
        main_df.loc[main_df['CT'] == "", 'CT'] = main_df['CT_temp']
        main_df.drop(columns=['CT_temp'], inplace=True)

        # Set column order
        main_df = main_df[['Argo ID','Slot ID/UTID','Ship Qtr','Ship Recog Qtr','Build Qtr','Ship Revenue Type',
                           'Build Product','Forecast ID','Sales Order','Fab Name','Committed Ship $','Region','IncoTerms',
                           'Holds','SO Status','Slot Request Date','MFG Commit Date','Ship Recog Date','MRP Date',
                           'Flex 02','Build Complete','PGI Date','PO expected date','PO date','SO date','CT',
                           'Configuration Note','Gate 2.7 plan','Gate 2.7 actual','Gate 3.5 plan','Gate 3.5 actual',
                           'Gate 5.5 plan','Gate 5.5 actual','Gate 6.5 plan','Gate 6.5 actual']]

        # ===== EXCEL FORMATTING =====
        wb = load_workbook(prev_file)
        ws = wb['Configuration tracking']

        border_style = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='thin', color='000000'))
        fill_blue = PatternFill(start_color='B8CCE4', end_color='c0ded9', fill_type='solid')
        fill_green = PatternFill(start_color='d9ecd0', end_color='d9ecd0', fill_type='solid')
        fill_pink = PatternFill(start_color='FCD1E0', end_color='FCD1E0', fill_type='solid')
        fill_orange = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

        def apply_common_style(ws, df, prev_df):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
                    cell.border = None
                    cell.font = None
                    cell.alignment = None
                    cell.fill = PatternFill(fill_type=None)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border_style
                    cell.font = Font(name='Calibri', size=10)
                    cell.alignment = Alignment(horizontal='center')

                    if r_idx == 1:
                        if c_idx in [24, 25]:
                            cell.fill = fill_pink
                        elif c_idx == 26:
                            cell.fill = fill_yellow
                        elif c_idx == 23:
                            cell.fill = fill_orange
                        elif c_idx in [29, 31, 33, 35, 37]:
                            cell.fill = fill_green
                        else:
                            cell.fill = fill_blue
                    else:
                        argo_id = row[0]
                        if argo_id not in prev_df['Argo ID'].values:
                            cell.fill = fill_yellow

        def adjust_column_widths(ws):
            for column in ws.columns:
                max_len = 0
                for cell in column:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[get_column_letter(column[0].column)].width = max_len + 2

        apply_common_style(ws, main_df, prev)
        adjust_column_widths(ws)

        # ===== FORMULAS =====
        for row in range(2, len(main_df) + 2):
            ws[f'AB{row}'] = f'=IF(ISBLANK(X{row}), "", X{row} + 7)'
            ws[f'AD{row}'] = f'=IF(ISNUMBER(AB{row}), AB{row} + 14, "")'
            ws[f'AF{row}'] = f'=IF(OR(ISBLANK(Z{row}), ISBLANK(S{row})), "", S{row} - Z{row} - 14)'
            ws[f'AH{row}'] = f'=IF(ISBLANK(S{row}), "", S{row} - 10)'

        # ===== EXPORT =====
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        today_date = datetime.datetime.today().strftime('%Y-%m-%d')
        new_file_name = f"Configuration_tracking_{today_date}.xlsx"

        st.success("File processed successfully and ready to download.")
        st.download_button(
            label="Download Updated File",
            data=output,
            file_name=new_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    st.info("Please upload all required files to continue.")
